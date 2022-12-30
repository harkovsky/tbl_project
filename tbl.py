import openpyxl


def autoxlsx():
    dest_filename = '202301 Табель ОГМ ПК 250.xlsx'
    wb = openpyxl.reader.excel.load_workbook(filename=dest_filename, data_only=True)
    wb.active = 0
    sheet = wb.active
    colour1 = 'BCE6FF'
    colour2 = 'BCFFF6'
    #colour2 = 'FFFFFF'
    colour3 = 'FFFFFF'

    # print(wb.sheetnames)
    # print(sheet['A8'].value)
    # sheet['A8'].value = '№ п/п'
    # ------clear results------

    for row in sheet['AJ13:AU41']:
        for cell in row:
            cell.value = None

    for row in sheet['E45:AI45']:
        for cell in row:
            if cell.value == 'В': cell.fill = openpyxl.styles.PatternFill(start_color=colour1, end_color=colour1, fill_type='solid')
            if cell.value == 'Р': cell.fill = openpyxl.styles.PatternFill(start_color=colour2, end_color=colour2, fill_type='solid')
            if cell.value == 'Р*': cell.fill = openpyxl.styles.PatternFill(start_color=colour2, end_color=colour2, fill_type='solid')
            if cell.value == 'х': cell.fill = openpyxl.styles.PatternFill(start_color=colour3, end_color=colour3, fill_type='solid')

    for rows in sheet['E13':'AI41']:
        komandirovka = 0
        sumhour = 0
        counthour = 0
        nochnye = 0
        aotp = 0
        otpusk = 0
        bolnichny = 0
        progul = 0
        prazdn = 0
        vsegoday = 0
        vsegohour = 0
        for cell in rows:
            cell.fill = openpyxl.styles.PatternFill(start_color=colour1, end_color=colour1, fill_type='solid')
            if str(cell.value).isdigit():
                sumhour += cell.value
                counthour += 1
                if sheet.cell(row=45, column=cell.column).value == 'В':
                    prazdn += cell.value
            if str(cell.value) == '4Н':
                sumhour += 4
                counthour += 1
                nochnye += 2
                if sheet.cell(row=45, column=cell.column).value == 'В':
                    prazdn += 4
            if str(cell.value) == '7Н':
                sumhour += 7
                counthour += 1
                nochnye += 6
                if sheet.cell(row=45, column=cell.column).value == 'В':
                    prazdn += 7
            if str(cell.value) == '8Н':
                sumhour += 8
                counthour += 1
                nochnye += 3.5
                if sheet.cell(row=45, column=cell.column).value == 'В':
                    prazdn += 8
            if str(cell.value) == '11Н':
                sumhour += 11
                counthour += 1
                nochnye += 8
                if sheet.cell(row=45, column=cell.column).value == 'В':
                    prazdn += 11
            if str(cell.value) == 'К': komandirovka += 1
            if str(cell.value) == 'А': aotp += 1
            if str(cell.value) == 'О': otpusk += 1
            if str(cell.value) == 'Б': bolnichny += 1
            if str(cell.value) == 'П': progul += 1
            if sheet.cell(row=45, column=cell.column).value == 'Р' and cell.value != '':
                cell.fill = openpyxl.styles.PatternFill(start_color=colour2, end_color=colour2, fill_type='solid')
                vsegoday += 1
                vsegohour += 8
            if sheet.cell(row=45, column=cell.column).value == 'Р*' and cell.value != '':
                cell.fill = openpyxl.styles.PatternFill(start_color=colour2, end_color=colour2, fill_type='solid')
                vsegoday += 1
                vsegohour += 7
            if sheet.cell(row=45, column=cell.column).value == 'х':
                cell.fill = openpyxl.styles.PatternFill(start_color=colour3, end_color=colour3, fill_type='solid')
        if komandirovka > 0: sheet.cell(cell.row, column=36).value = komandirovka
        if sumhour > 0: sheet.cell(cell.row, column=45).value = sumhour
        if counthour > 0: sheet.cell(cell.row, column=44).value = counthour
        if nochnye > 0: sheet.cell(cell.row, column=42).value = nochnye
        if aotp > 0: sheet.cell(cell.row, column=40).value = aotp
        if otpusk > 0: sheet.cell(cell.row, column=39).value = otpusk
        if bolnichny > 0: sheet.cell(cell.row, column=38).value = bolnichny
        if progul > 0: sheet.cell(cell.row, column=37).value = progul
        if prazdn > 0: sheet.cell(cell.row, column=41).value = prazdn
        if counthour > 0: sheet.cell(cell.row, column=46).value = vsegoday
        if sumhour > 0: sheet.cell(cell.row, column=47).value = vsegohour
        if sumhour > vsegohour: sheet.cell(cell.row, column=43).value = sumhour - vsegohour

    # Итоговые результаты
    sheet['AR42'].value = '=SUM(AR13:AR36)'
    sheet['AT42'].value = '=SUM(AT13:AT36)'
    sheet['AS43'].value = '=SUM(AS13:AS36)'
    sheet['AU43'].value = '=SUM(AU13:AU36)'
    wb.save(dest_filename)


if __name__ == '__main__':
    autoxlsx()
